"""
------------------------------------------------------------
ECAT
Excel Exporter
Build : 1.1.7
------------------------------------------------------------
"""

from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side
)
from openpyxl.utils import get_column_letter


class ExcelExporter:

    def __init__(self):

        self.title_fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78"
        )

        self.header_fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAD3"
        )

        self.zone_fill = PatternFill(
            fill_type="solid",
            fgColor="FFF2CC"
        )

        self.division_fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78"
        )

        self.bold = Font(bold=True)

        self.white_bold = Font(
            bold=True,
            color="FFFFFF"
        )

        self.center = Alignment(
            horizontal="center",
            vertical="center"
        )

        self.right = Alignment(
            horizontal="right"
        )

        self.left = Alignment(
            horizontal="left"
        )

        thin = Side(style="thin")

        self.border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin
        )

    # --------------------------------------------------------

    def export(

        self,

        dataframe,

        filename,

        report_title,

        period="",

        filters=""

    ):

        wb = Workbook()

        ws = wb.active

        ws.title = "Report"

        total_columns = len(dataframe.columns)

        # -----------------------------------------------
        # Report Title
        # -----------------------------------------------

        ws.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=total_columns
        )

        cell = ws.cell(
            row=1,
            column=1
        )

        cell.value = report_title

        cell.font = self.white_bold

        cell.fill = self.title_fill

        cell.alignment = self.center

        # -----------------------------------------------
        # Period
        # -----------------------------------------------

        ws.cell(
            row=2,
            column=1
        ).value = f"Period : {period}"

        # -----------------------------------------------
        # Generated On
        # -----------------------------------------------

        ws.cell(
            row=3,
            column=1
        ).value = (
            "Generated On : "
            + datetime.now().strftime(
                "%d-%b-%Y %I:%M %p"
            )
        )

        # -----------------------------------------------
        # Filters
        # -----------------------------------------------

        ws.cell(
            row=4,
            column=1
        ).value = f"Filters : {filters}"

        # -----------------------------------------------
        # Header
        # -----------------------------------------------

        start_row = 6

        for col, name in enumerate(
            dataframe.columns,
            start=1
        ):

            cell = ws.cell(
                row=start_row,
                column=col
            )

            cell.value = name

            cell.font = self.bold

            cell.fill = self.header_fill

            cell.alignment = self.center

            cell.border = self.border

        # -----------------------------------------------
        # Data
        # -----------------------------------------------

        row_no = start_row + 1

        for row in dataframe.itertuples(index=False):

            for col_no, value in enumerate(
                row,
                start=1
            ):

                cell = ws.cell(
                    row=row_no,
                    column=col_no
                )

                cell.value = value

                cell.border = self.border

                if col_no == 1:
                    cell.alignment = self.left
                else:
                    cell.alignment = self.right

                    if isinstance(
                        value,
                        (int, float)
                    ):
                        cell.number_format = '#,##0'

            feeder = str(row[0]).upper()

            if "ZONE" in feeder:

                for c in range(
                    1,
                    total_columns + 1
                ):

                    ws.cell(
                        row=row_no,
                        column=c
                    ).fill = self.zone_fill

                    ws.cell(
                        row=row_no,
                        column=c
                    ).font = self.bold

            if "DIVISION" in feeder:

                for c in range(
                    1,
                    total_columns + 1
                ):

                    ws.cell(
                        row=row_no,
                        column=c
                    ).fill = self.division_fill

                    ws.cell(
                        row=row_no,
                        column=c
                    ).font = self.white_bold

            row_no += 1

        # -----------------------------------------------
        # Auto Width
        # -----------------------------------------------

        for column in ws.columns:

            length = 0

            letter = get_column_letter(
                column[0].column
            )

            for cell in column:

                try:

                    length = max(
                        length,
                        len(str(cell.value))
                    )

                except:

                    pass

            ws.column_dimensions[
                letter
            ].width = min(
                length + 3,
                35
            )

        # -----------------------------------------------
        # Freeze Header
        # -----------------------------------------------

        ws.freeze_panes = "A7"

        wb.save(filename)