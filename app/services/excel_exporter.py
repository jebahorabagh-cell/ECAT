"""
------------------------------------------------------------
ECAT
Excel Exporter
Build : 1.2.3
------------------------------------------------------------
"""

from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment


class ExcelExporter:

    def export(self, dataframe, report_name="Report"):

        workbook = openpyxl.Workbook()

        sheet = workbook.active

        sheet.title = "Report"

        self._write_dataframe(sheet, dataframe)

        self._format(sheet)

        from pathlib import Path

        # User's Downloads folder
        export_folder = Path.home() / "Downloads"

        filename = (
            f"{report_name}_"
            f"{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        )

        filepath = export_folder / filename

        workbook.save(filepath)

        return filepath
        
    def _write_dataframe(self, sheet, df):

        # Header

        for col, name in enumerate(df.columns, start=1):

            sheet.cell(
                row=1,
                column=col,
                value=name
            )

        # Data

        for r, row in enumerate(df.itertuples(index=False), start=2):

            for c, value in enumerate(row, start=1):

                sheet.cell(
                    row=r,
                    column=c,
                    value=value
                )
                
    def _format(self, sheet):

        # Header

        header_fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAD3"
        )

        for cell in sheet[1]:

            cell.font = Font(bold=True)

            cell.fill = header_fill

            cell.alignment = Alignment(horizontal="center")

        # Freeze header

        sheet.freeze_panes = "A2"

        # Auto Filter

        sheet.auto_filter.ref = sheet.dimensions

        # Auto Width

        for column in sheet.columns:

            width = 15

            letter = column[0].column_letter

            for cell in column:

                if cell.value:

                    width = max(width, len(str(cell.value)) + 2)

            sheet.column_dimensions[letter].width = width